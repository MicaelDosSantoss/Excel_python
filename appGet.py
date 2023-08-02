import requests
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill


book = openpyxl.Workbook() # Criação da planilha

# Visualização de paginas do excel

print(book.sheetnames)

# Criar pagina 
book.create_sheet('Pokedex') 

pokedex_page = book['Pokedex'] #Pagína selecionada 

pokedex_nacional = 'https://pokeapi.co/api/v2/pokedex/1/' # API publica da Pókedex

get = requests.get(pokedex_nacional)  # GET para pegar a API
Pokedex_json = get.json() # Transformando o resultado do GET em JSON

bold_number = Font(bold=True) # Número em negrito

title = ['Número Pokedex','Pokemon'] # Titulo do Excel

# Ajustar a altura da linha para acomodar o conteúdo
pokedex_page.row_dimensions[1].height = 40
pokedex_page.row_dimensions[1].width = 40


pokedex_page.append(title)  # Inserção do Titulo na tabela

cont = 1  # quantidade de itens para descobrir o limite

for pokemon in Pokedex_json['pokemon_entries']: # Para pegar os itens, um por um, selecionado apenas os nomes

    namePoke = pokemon["pokemon_species"]['name']   # Variavel para armazernar

    pokedex_page.append([pokemon["entry_number"], namePoke]) # Inserção de número da Pokedex é os nomes dos pokemons

    cont += 1 

str(cont) # Transformar em sting

Align_center = Alignment(horizontal='center',vertical='center') # Alinhar os itens nas células do excel

border_style = Side(style='thin') # Inserir bordar na tabela

for row in pokedex_page.iter_rows(min_row=1, max_row=cont,min_col=1,max_col=len(title) ): 
    for cell in row:
        cell.alignment = Align_center
        cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

for row in pokedex_page.iter_rows(min_col=1, max_row=1, min_row=1, max_col=len(title)):
    for cell in row:
        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        cell.font = Font(size=15)

for row in pokedex_page.iter_rows(min_row=1, max_row=cont, min_col=1, max_col=1):
    for cell in row:
        cell.font = Font(bold=True)


book.save('Pokedex.xlsx') # Salvar as alterações