import openpyxl
import json

workbook = openpyxl.load_workbook('Pokedex.xlsx') # carregar o arquivo

Pokedex = workbook.sheetnames[1] # Selecionando a Planilha
sheet = workbook[Pokedex] 

Pokedex_list = [] # Lista que vai receber o valores

for row in sheet.iter_rows(values_only=True): # Nessa linha estou percorrendo todo todas as linhas da tabela, o values_only = True indica que eu deseja obter apenas os valores das células em vez de objetos de células completos
    Pokedex_list.append([{
        'numero pokedex': row[0],
        'Pokemon': row[1]
    }])


with open('Pokedex.json', 'w') as pokedex_json: # Criação do arquivo JSON através do codigo
    json.dump(Pokedex_list, pokedex_json, indent=2)
